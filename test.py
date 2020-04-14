import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
"""
from openpyxl.workbook import Workbook
from openpyxl.chart import PieChart
from openpyxl import load_workbook
"""

all_data = pd.read_csv('data/SpeCopyspec.csv')
print("\n info about the all data \n")

print("number of scores ", len(all_data['Score'].unique()))
print(all_data.info())
numerics = ['int16', 'int32', 'int64', 'float16', 'float32', 'float64']

newdf = all_data.select_dtypes(include=numerics)
print(newdf.columns)
# example
""" 
wb = load_workbook('SpeCopy.xlsx')
ws = wb.active
cell_range = ws['A2':'C2']
col_range = ws['A' : 'C']
row_range =ws[1:5]
for row in ws.iter_rows(min_row=1, max_col=20, max_row=20, values_only=True):
    for cell in row:        
        if   cell == 'DIT400TR-55R/55RL':
            print(cell)

all_data = pd.read_csv('SpeCopyspec.csv')
print("\n info about the all data \n")
print(all_data.info())

# drop all except vendor , model , score
data_filter = all_data.filter(['Vendor','Model', 'Score'])
print("\n info about the filtered data \n")
print(data_filter.info())

print("\n check filtered data head \n")
print(data_filter.shape)




print("\n find max  and min \n")
data_max = data_filter.groupby(['Vendor','Model'], sort=False)['Score'].transform(max) == data_filter['Score']
data_min = data_filter.groupby(['Vendor', 'Model'], sort=False)['Score'].transform(min) == data_filter['Score']

data_max = data_filter[data_max].copy()


x_data_max = np.arange(data_max['Vendor'].shape[0])

plt.barh(x_data_max, data_max["Score"])
plt.yticks(x_data_max, data_max['Vendor'], fontsize=5)
plt.xlabel('Scores')
plt.ylabel('Vendors & Models')
plt.show()




"""